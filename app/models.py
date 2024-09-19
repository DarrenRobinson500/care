from django.db.models import *
from datetime import datetime, timedelta
import pytz
from django.utils import timezone
from django.conf import settings
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .sms import *
import pandas as pd

one_week_ago = timezone.now() - timedelta(days=7)
MONTH_FORMAT = '%B %Y'
DATE_FORMAT = '%a, %d %b'
DATETIME_FORMAT = '%d %b %H:%M'
def get_choices(choices):
    result = []
    for choice in choices: result.append(((choice), (choice)))
    return result

class Patient(Model):
    model_str = "patient"
    single = "Patient"
    plural = "Patients"
    has_order = True
    name = TextField(max_length=250, null=True, blank=True)
    order = IntegerField(null=True, blank=True)
    active = BooleanField(null=True, blank=True, default=True)
    mobile = TextField(null=True, blank=True)
    photo = ImageField(null=True, blank=True, upload_to="images/")
    def __str__(self): return self.name
    def notes(self): return Note.objects.filter(patient=self)
    def info(self): return Info.objects.filter(patient=self).order_by('order_field')
    def medications(self): return Medication.objects.filter(patient=self).order_by('order')
    def jobs(self): return Job.objects.filter(patient=self)
    def jobs_open(self): return Job.objects.filter(patient=self, date_time_completed__isnull=True)
    def jobs_complete(self): return Job.objects.filter(patient=self, date_time_completed__isnull=False)
    def update_links(self):
        info_fields = InfoField.objects.filter(database="patient")
        infos = self.info()
        for info_field in info_fields:
            found = False
            for info in infos:
                if info.field == info_field.field: found = True
            if not found:
                new_info = Info(category=info_field.category, field=info_field.field, order_category=info_field.category.order, order_field=info_field.order,
                                info_category=info_field.category, info_field=info_field)
                new_info.patient = self
                new_info.save()
        for info in self.info():
            if not info.info_field:
                info.delete()

    def recurring_jobs(self):
        result = []
        jobtypes = JobType.objects.all()
        for jobtype in jobtypes:
            if jobtype.recurring:
                existing = RecurringJob.objects.filter(patient=self, jobtype=jobtype).first()
                result.append((jobtype, existing))
        return result
    def recurring_jobs_requested(self):
        result = []
        jobtypes = JobType.objects.all()
        for jobtype in jobtypes:
            if jobtype.recurring:
                existing = RecurringJob.objects.filter(patient=self, jobtype=jobtype).first()
                if existing:
                    result.append((jobtype, existing))
        return result
    def recurring_jobs_not_requested(self):
        result = []
        jobtypes = JobType.objects.all()
        for jobtype in jobtypes:
            existing = RecurringJob.objects.filter(patient=self, jobtype=jobtype).first()
            if not existing and jobtype.recurring:
                result.append((jobtype, existing))
        return result
    def financials_current_month(self):
        return self.financials(8, 2024)
    def financials(self, month, yeat):
        jobs = self.jobs_complete().filter(date_time_completed__month=month)
        category_counts = jobs.values('jobtype').annotate(count=Count('jobtype'))
        job_list = []
        for category in category_counts:
            jobtype = JobType.objects.get(id=category['jobtype'])
            count = category['count']
            unit_cost = jobtype.amount
            jobtype_cost = count * unit_cost
            string = f"{jobtype.name}: {count} x {unit_cost} = {jobtype_cost}"
            job_list.append((jobtype, count, unit_cost, jobtype_cost, string))
        return job_list

# STAFF_CHOICES = (('Staff', 'Staff'), ('Admin', 'Admin'), ('Setup', 'Setup'))

STAFF_CHOICES = get_choices(["Staff", "Admin", "Setup"])
colours = ["#50CDE6", "#FC98CC", "#E6CD50", "#95C74E", "#FFB239", "#F7C9C9"]

class Staff(Model):
    model_str = "staff"
    single = "Staff Member"
    plural = "Staff"
    has_order = True
    name = TextField(max_length=250, null=True, blank=True)
    first_name = TextField(max_length=250, null=True, blank=True)
    last_name = TextField(max_length=250, null=True, blank=True)
    manager = ForeignKey('self', null=True, blank=True, on_delete=SET_NULL)
    order = IntegerField(null=True, blank=True)
    colour_no = IntegerField(null=True, blank=True, default=0)
    active = BooleanField(null=True, blank=True, default=True)
    user = ForeignKey(settings.AUTH_USER_MODEL, null=True, blank=True, on_delete=CASCADE)
    user_type = TextField(null=True, blank=True, choices=STAFF_CHOICES)
    mobile = TextField(null=True, blank=True)
    photo = ImageField(null=True, blank=True, upload_to="images/")
    def __str__(self): return f"{self.name}"
    def valid_mobile(self):
        if self.mobile: return len(self.mobile) == 10
    def colour(self):
        # print("Colour:", self.id, colours[self.colour_no])
        return colours[min(self.colour_no, len(colours) - 1)]
    def increment_colour(self):
        print("Pre:", self.id, self.colour_no, self.colour())
        self.colour_no += 1
        # if self.colour_no > len(colours): self.colour_no = 0
        self.save()
        print("Post:", self.id, self.colour_no, self.colour())
    def notes(self): return Note.objects.filter(staff=self)
    def info(self): return Info.objects.filter(staff=self).order_by('order_field').order_by('order_category')
    def open_jobs(self): return Job.objects.filter(staff=self, date_time_completed__isnull=True)
    def recent_jobs(self): return Job.objects.filter(staff=self, date_time_completed__gte=one_week_ago)
    def on_duty(self):
        now = datetime.now().replace(tzinfo=None)
        shifts = self.shifts()
        for shift in shifts:
            if shift.start.replace(tzinfo=None) <= now <= shift.end.replace(tzinfo=None): return True
        return False
    def next_fortnight(self):
        result = []
        day = datetime.now().date()
        available_shifts = AvailableShift.objects.filter(active=True).order_by('order')
        shifts = Shift.objects.filter(staff=self, date__gte=day, date__lte=day+timedelta(days=15))
        for x in range(15):
            shift_list = []
            for shift in available_shifts:
                shift_list.append((shift.id, shifts.filter(date=day, shift=shift).exists()))
            result_day = [day, x, shift_list]
            result.append(result_day)
            day = day + timedelta(days=1)
        return result
    def shifts(self): return Shift.objects.filter(staff=self)
    def available_shifts(self): return AvailableShift.objects.filter(active=True).order_by('order')
    def preferred_shifts(self): return PreferredShift.objects.filter(staff=self)

    def update_links(self):
        # print("Updating links")
        info_fields = InfoField.objects.filter(database="staff")
        infos = self.info()
        for info_field in info_fields:
            found = False
            for info in infos:
                if info.field == info_field.field: found = True
            if not found:
                new_info = Info(category=info_field.category, field=info_field.field, order_category=info_field.category.order, order_field=info_field.order,
                                info_category=info_field.category, info_field=info_field)
                new_info.staff = self
                new_info.save()
        for info in self.info():
            if not info.info_field:
                info.delete()
        shifts = AvailableShift.objects.all()
        preferred_shifts = self.preferred_shifts()
        for shift in shifts:
            found = False
            for preferred_shift in preferred_shifts:
                if preferred_shift.available_shift == shift: found = True
            if not found:
                new_item = PreferredShift(staff=self, available_shift=shift)
                new_item.save()
            # print("Updating links:", shift, found)

class Day(Model):
    model_str = "day"
    single = "Roster"
    plural = single
    has_order = False
    day = DateField(null=True, blank=True)
    def __str__(self): return self.day.strftime(DATE_FORMAT)
    def previous(self): return self.day_adj(-1)
    def next(self): return self.day_adj(1)
    def day_adj(self, adj): return Day.objects.filter(day=self.day + timedelta(days=adj)).first()
    def initiate(self): self.make_shifts()
    def make_shifts(self):
        for staff in Staff.objects.all():
            for preferred_shift in staff.preferred_shifts():
                preferred_shift.create_shift(self.day)
    def working_hours(self):
        staff_hours = Shift.objects.filter(date=self.day).aggregate(total=Sum('duration'))['total']
        if not staff_hours: staff_hours = 0
        return round(staff_hours, 1)
    def shifts(self):
        return Shift.objects.filter(date=self.day)
    def recurring_jobs(self):
        day_of_week = self.day.strftime("%A")
        if day_of_week == "Sunday": recurring_jobs = RecurringJob.objects.filter(sunday=True)
        if day_of_week == "Monday": recurring_jobs = RecurringJob.objects.filter(monday=True)
        if day_of_week == "Tuesday": recurring_jobs = RecurringJob.objects.filter(tuesday=True)
        if day_of_week == "Wednesday": recurring_jobs = RecurringJob.objects.filter(wednesday=True)
        if day_of_week == "Thursday": recurring_jobs = RecurringJob.objects.filter(thursday=True)
        if day_of_week == "Friday": recurring_jobs = RecurringJob.objects.filter(friday=True)
        if day_of_week == "Saturday": recurring_jobs = RecurringJob.objects.filter(saturday=True)
        return recurring_jobs
    def job_hours(self):
        recurring_jobs = self.recurring_jobs()
        try:
            return round(recurring_jobs.aggregate(total=Sum('duration'))['total'] / 60, 1)
        except:
            return 0
    def working_over_jobs(self):
        working = self.working_hours()
        if not working: working = 0
        jobs = self.job_hours()
        if jobs != 0:
            return f"{int(working / jobs * 100)}%"
        return ""

class Month(Model):
    model_str = "month"
    single = "Invoice"
    plural = single + "s"
    has_order = False
    month = IntegerField(null=True, blank=True)
    year = IntegerField(null=True, blank=True)
    def __str__(self):
        date = datetime(self.year, self.month, 1)
        return date.strftime(MONTH_FORMAT)
    def patients(self): return Patient.objects.all()
    def patient_data(self):
        data = []
        for patient in self.patients():
            data.append((patient, patient.financials(self.month, self.year)))
        return data
    def create_financial_spreadsheet(self):
        wb = Workbook()
        ws = wb.active
        ws.title = str(self)
        row = 1
        for column, heading in enumerate(["Patient", "Job", "Count", "Unit Cost", "Cost"], 1):
            cell = ws.cell(row=row, column=column, value=heading)
            cell.font = Font(bold=True)
            if column >= 3: cell.alignment = Alignment(horizontal='right')
        row = 2
        for patient in self.patients():
            for job, count, unit_cost, cost, string in patient.financials(self.month, self.year):
                ws.cell(row=row, column=1, value=str(patient))
                ws.cell(row=row, column=2, value=str(job))
                ws.cell(row=row, column=3, value=count)
                ws.cell(row=row, column=4, value=unit_cost)
                ws.cell(row=row, column=5, value=cost)
                row += 1
        return wb

class AvailableShift(Model):
    model_str = "availableshift"
    single = "Available Shift"
    plural = single + "s"
    has_order = True
    start = TimeField(null=True, blank=True)
    end = TimeField(null=True, blank=True)
    order = IntegerField(null=True, blank=True)
    active = BooleanField(null=True, blank=True, default=False)
    def __str__(self): return f"{self.start} - {self.end}"

class PreferredShift(Model):
    model_str = "preferredshift"
    single = "Preferred Shift"
    plural = single + "s"
    has_order = True
    available_shift = ForeignKey(AvailableShift, null=True, blank=True, on_delete=CASCADE)
    staff = ForeignKey(Staff, null=True, blank=True, on_delete=CASCADE)
    sunday = BooleanField(null=True, blank=True, default=False)
    monday = BooleanField(null=True, blank=True, default=False)
    tuesday = BooleanField(null=True, blank=True, default=False)
    wednesday = BooleanField(null=True, blank=True, default=False)
    thursday = BooleanField(null=True, blank=True, default=False)
    friday = BooleanField(null=True, blank=True, default=False)
    saturday = BooleanField(null=True, blank=True, default=False)
    order = IntegerField(null=True, blank=True)
    def __str__(self): return f"{self.staff} {self.available_shift}"
    def preference_by_day(self):
        return ("sunday", self.sunday), ("monday", self.monday), ("tuesday", self.tuesday), ("wednesday", self.wednesday), \
            ("thursday", self.thursday), ("friday", self.friday), ("saturday", self.saturday)
    def create_shift(self, date_obj):
        day = date_obj.strftime("%A")
        make_shift = False

        if day == "Sunday" and self.sunday: make_shift = True
        if day == "Monday" and self.monday: make_shift = True
        if day == "Tuesday" and self.tuesday: make_shift = True
        if day == "Wednesday" and self.wednesday: make_shift = True
        if day == "Thursday" and self.thursday: make_shift = True
        if day == "Friday" and self.friday: make_shift = True
        if day == "Saturday" and self.saturday: make_shift = True


        print("Make shifts", day == "Tuesday", day == "Wednesday", make_shift)

        if make_shift:
            existing = Shift.objects.filter(staff=self.staff, shift=self.available_shift, date=date_obj)
            print("Existing shift:", existing)
            if len(existing) == 0:
                print("Create Shift", self.staff, str(date_obj))
                new = Shift(staff=self.staff, shift=self.available_shift, date=date_obj)
                new.save()
                new.save_start_and_end()

class Shift(Model):
    model_str = "shift"
    single = "Shift"
    plural = single + "s"
    has_order = True
    staff = ForeignKey(Staff, null=True, blank=True, on_delete=CASCADE)
    shift = ForeignKey(AvailableShift, null=True, blank=True, on_delete=SET_NULL)
    start = DateTimeField(null=True, blank=True)
    end = DateTimeField(null=True, blank=True)
    duration = FloatField(null=True, blank=True)
    date = DateField(blank=True, null=True)
    order = IntegerField(null=True, blank=True)
    notes = TextField(max_length=2500, null=True, blank=True)
    def __str__(self): return f"{self.staff} {self.date} {self.shift}"
    def save_start_and_end(self):
        self.start = datetime.combine(self.date, self.shift.start)
        self.end = datetime.combine(self.date, self.shift.end)
        if self.end < self.start: self.end = self.end + timedelta(days=1)
        self.duration = round((self.end - self.start).total_seconds() / 3600, 1)
        print("Shift:", self.duration, self.end, self.start)
        self.save()

class JobType(Model):
    model_str = "jobtype"
    single = "Job Type"
    plural = "Job Types"
    has_order = True
    name = TextField(max_length=250, null=True, blank=True)
    amount = IntegerField(null=True, blank=True)
    order = IntegerField(null=True, blank=True)
    medication = BooleanField(null=True, blank=True, default=False)
    active = BooleanField(null=True, blank=True, default=True)
    recurring = BooleanField(null=True, blank=True, default=True)
    duration = IntegerField(null=True, blank=True)
    start = TimeField(null=True, blank=True)
    end = TimeField(null=True, blank=True)
    def __str__(self): return self.name
    def times(self): return f"{self.start} to {self.end}"

class Frequency(Model):
    model_str = "frequency"
    single = "Frequency"
    plural = "Frequencies"
    has_order = True
    order = IntegerField(null=True, blank=True)
    name = TextField(max_length=250, null=True, blank=True)
    sunday = BooleanField(null=True, blank=True, default=False)
    monday = BooleanField(null=True, blank=True, default=False)
    tuesday = BooleanField(null=True, blank=True, default=False)
    wednesday = BooleanField(null=True, blank=True, default=False)
    thursday = BooleanField(null=True, blank=True, default=False)
    friday = BooleanField(null=True, blank=True, default=False)
    saturday = BooleanField(null=True, blank=True, default=False)
    per_day = IntegerField(null=True, blank=True, default=1)
    def __str__(self): return self.name

class RecurringJob(Model):
    model_str = "recurringjob"
    single = "Regular Job"
    plural = "Regular Jobs"
    has_order = False
    jobtype = ForeignKey(JobType, null=True, blank=True, on_delete=SET_NULL)
    patient = ForeignKey(Patient, null=True, blank=True, on_delete=SET_NULL)
    frequency = ForeignKey(Frequency, null=True, blank=True, on_delete=SET_NULL)
    sunday = BooleanField(null=True, blank=True, default=False)
    monday = BooleanField(null=True, blank=True, default=False)
    tuesday = BooleanField(null=True, blank=True, default=False)
    wednesday = BooleanField(null=True, blank=True, default=False)
    thursday = BooleanField(null=True, blank=True, default=False)
    friday = BooleanField(null=True, blank=True, default=False)
    saturday = BooleanField(null=True, blank=True, default=False)
    duration = IntegerField(null=True, blank=True)
    def __str__(self): return f"{self.jobtype} for {self.patient}"
    def duration_hours(self): return str(round(self.duration / 60, 2))
    def save_initial(self):
        freq = self.frequency
        self.sunday = freq.sunday
        self.monday = freq.monday
        self.tuesday = freq.tuesday
        self.wednesday = freq.wednesday
        self.thursday = freq.thursday
        self.friday = freq.friday
        self.saturday = freq.saturday
        self.duration = self.jobtype.duration
        self.save()
    def notes(self): return Note.objects.filter(recurring_job=self)

class Job(Model):
    model_str = "job"
    single = "Job"
    plural = "Jobs"
    has_order = False
    jobtype = ForeignKey(JobType, null=True, blank=True, on_delete=SET_NULL)
    patient = ForeignKey(Patient, null=True, blank=True, on_delete=SET_NULL)
    staff = ForeignKey(Staff, null=True, blank=True, on_delete=SET_NULL)
    date_time_requested = DateTimeField(null=True, blank=True)
    date_time_completed = DateTimeField(null=True, blank=True)
    notes = TextField(max_length=2500, null=True, blank=True)
    amount = IntegerField(null=True, blank=True)
    start = DateTimeField(null=True, blank=True)
    end = DateTimeField(null=True, blank=True)
    def __str__(self):
        return f"{self.jobtype} for {self.patient}"
    def times(self):
        if self.start and self.end:
            return f"{self.start.strftime(DATETIME_FORMAT)} to {self.end.strftime(DATETIME_FORMAT)}"
    def available_staff(self):
        available_staff = []
        shifts = Shift.objects.all()
        for shift in shifts:
            if shift.staff not in available_staff:
                if shift.start and self.start and shift.end:
                    if shift.start < self.start < shift.end:
                         available_staff.append(shift.staff)
            if shift.staff not in available_staff:
                if shift.start and self.end and shift.end:
                    print("Available staff:", shift.start, self.end, shift.end, shift.start < self.end < shift.end)
                    if shift.start < self.end < shift.end:
                        available_staff.append(shift.staff)
        return available_staff

class MedicationType(Model):
    model_str = "medicationtype"
    single = "Medication Type"
    plural = "Medication Types"
    has_order = True
    name = TextField(max_length=250, null=True, blank=True)
    category = TextField(max_length=250, null=True, blank=True)
    order = IntegerField(null=True, blank=True)
    def __str__(self): return f"{self.name} ({self.category})"

medication_frequency = get_choices(["Daily", "Twice a Day", "Every Second Day"])
medication_timing = get_choices(["Morning", "Evening", "Morning and Evening", "Anytime"])

class Medication(Model):
    model_str = "medication"
    single = "Medication"
    plural = "Medications"
    has_order = True
    patient = ForeignKey(Patient, null=True, blank=True, on_delete=CASCADE)
    medication_type = ForeignKey(MedicationType, null=True, blank=True, on_delete=SET_NULL)
    dosage = TextField(max_length=250, null=True, blank=True)
    frequency = ForeignKey(Frequency, null=True, blank=True, on_delete=SET_NULL)
    # frequency = TextField(null=True, blank=True)
    first_day = DateField(null=True, blank=True)
    order = IntegerField(null=True, blank=True)
    def __str__(self): return self.name
    def initiate(self, id):
        self.patient = Patient.objects.get(id=id)
        self.first_day = datetime.today()
        self.save()

class InfoCategory(Model):
    model_str = "infocategory"
    single = "Info Category"
    plural = "Info Categories"
    has_order = True
    category = TextField(null=True, blank=True)
    order = IntegerField(null=True, blank=True)
    def __str__(self): return self.category
    def parent_model(self):
        if self.database == "patient": return Patient
        if self.database == "staff": return Staff
        if self.database == "recurring_job": return RecurringJob
        if self.database == "job": return Job
        if self.database == "month": return Month

class InfoField(Model):
    model_str = "infofield"
    single = "Info Field"
    plural = "Info Fields"
    has_order = True
    database = TextField(null=True, blank=True, choices=get_choices(["patient", "staff", "recurring_job", "job", "month"]))
    category = ForeignKey(InfoCategory, null=True, blank=True, on_delete=SET_NULL)
    field = TextField(null=True, blank=True)
    order = IntegerField(null=True, blank=True)
    data_type = TextField(null=True, blank=True, choices=get_choices(["text", "text_area", "date", "number"]))
    def __str__(self):
        try:
            return self.field
        except:
            return "Error with Info Field"
    def parent_model(self):
        if self.database == "patient": return Patient
        if self.database == "staff": return Staff
        if self.database == "recurring_job": return RecurringJob
        if self.database == "job": return Job
        if self.database == "month": return Month
    def infos(self):
        return Info.objects.filter(info_field=self)
    def update_infos(self):
        for info in self.infos():
            info.order_field = self.order
            info.save()
            print("Updating:", self, info, self.order, info.order_field)

class Info(Model):
    model_str = "info"
    single = "Information"
    plural = "Information"
    has_order = False
    category = TextField(blank=True, null=True)
    field = TextField(blank=True, null=True)
    content_text = TextField(blank=True, null=True)
    content_date = DateField(null=True, blank=True)
    content_number = FloatField(null=True, blank=True)
    info_category = ForeignKey(InfoCategory, blank=True, null=True, related_name="info_type", on_delete=CASCADE)
    order_category = IntegerField(null=True, blank=True)
    info_field = ForeignKey(InfoField, blank=True, null=True, related_name="info_type", on_delete=CASCADE)
    order_field = IntegerField(null=True, blank=True)

    created_by = ForeignKey(Staff, blank=True, null=True, related_name="info_created_by", on_delete=SET_NULL)
    date_time_created = DateTimeField(null=True, blank=True)
    staff = ForeignKey(Staff, blank=True, null=True, related_name="info_staff", on_delete=SET_NULL)
    patient = ForeignKey(Patient, blank=True, null=True, related_name="info_patient", on_delete=SET_NULL)
    recurring_job = ForeignKey(RecurringJob, blank=True, related_name="info_recurring_job", null=True, on_delete=SET_NULL)
    job = ForeignKey(Job, blank=True, null=True, related_name="info_job", on_delete=SET_NULL)
    month = ForeignKey(Month, blank=True, null=True, related_name="info_month", on_delete=SET_NULL)
    def __str__(self):
        if self.content_text: return self.content_text
        if self.content_date: return str(self.content_date)
        if self.content_number: return str(self.content_number)
        return self.info_field.field
    def date_string(self): return self.date_time_created.strftime(DATE_FORMAT)
    def content(self):
        if not self.info_field: return None
        content_type = self.info_field.data_type
        # print("Content type:", content_type)
        if content_type in ["text", "text_area"]: return self.content_text
        if content_type == "date": return self.content_date
        if content_type == "number": return self.content_number
        return f"Error with content type"
    def save_content(self, value):
        content_type = self.info_field.data_type
        if content_type in ["text", "text_area"]: self.content_text = value
        if content_type == "date": self.content_date = value
        if content_type == "number": self.content_number = value
        self.save()

    def parent(self):
        if self.staff: return self.staff
        if self.patient: return self.patient
        if self.recurring_job: return self.recurring_job
        if self.job: return self.job
        if self.month: return self.month

    def parent_model(self): return self.info_field.parent_model()
    def parent_model_str(self): return self.info_field.parent_model().model_str
    def parent_str(self): return self.parent_model().single + ": " + str(self.parent())

class Note(Model):
    model_str = "note"
    single = "Note"
    plural = "Notes"
    has_order = False
    heading = TextField(blank=True, null=True)
    content = TextField(blank=True, null=True)
    created_by = ForeignKey(Staff, blank=True, null=True, related_name="note_created_by", on_delete=SET_NULL)
    date_time_created = DateTimeField(null=True, blank=True)
    staff = ForeignKey(Staff, blank=True, null=True, related_name="note_staff", on_delete=SET_NULL)
    patient = ForeignKey(Patient, blank=True, null=True, related_name="note_patient", on_delete=SET_NULL)
    recurring_job = ForeignKey(RecurringJob, blank=True, null=True, related_name="note_recurring_job", on_delete=SET_NULL)
    job = ForeignKey(Job, blank=True, null=True, related_name="note_job", on_delete=SET_NULL)
    month = ForeignKey(Month, blank=True, null=True, related_name="note_month", on_delete=SET_NULL)
    def __str__(self): return self.heading
    def date_string(self): return self.date_time_created.strftime(DATE_FORMAT)
    def parent(self):
        if self.staff: return self.staff
        if self.patient: return self.patient
        if self.recurring_job: return self.recurring_job

    def parent_model_str(self):
        if self.patient: return "patient"
        if self.staff: return "staff"
        if self.recurring_job: return "recurringjob"

    def parent_str(self):
        if self.patient: return "Patient: " + str(self.parent())
        if self.staff: return "Staff: " + str(self.parent())
        if self.recurring_job: return "Recurring Job: " + str(self.parent())

class SMS(Model):
    model_str = "sms"
    single = "SMS"
    plural = "SMSs"
    has_order = False
    staff = ForeignKey(Staff, blank=True, null=True, related_name="sms_staff", on_delete=SET_NULL)
    patient = ForeignKey(Patient, blank=True, null=True, related_name="sms_patient", on_delete=SET_NULL)
    sender = ForeignKey(Staff, blank=True, null=True, related_name="sms_sender", on_delete=SET_NULL)
    content = TextField(blank=True, null=True)
    date_time_completed = DateTimeField(null=True, blank=True)
    def __str__(self):
        if self.staff:
            return f"{self.staff} {self.content}"
        if self.patient:
            return f"{self.staff} {self.content}"
        return "Faulty SMS"
    def send(self):
        print("SMS Send", self.content, self.staff.mobile)
        if self.content and self.staff.mobile:
            to_mobile = "+61" + self.staff.mobile[1:]
            client.messages.create(body=self.content, from_=my_twilio_number, to=to_mobile)
            self.date_time_completed = datetime.now()
            self.sender = self.sender
            print("SMS Send:", self.staff.mobile, self.content)
            self.save()


file_types = ["patient", "staff", "shift", "infocategory", "infofield", "medicationtype", "jobtype", "frequency", "availableshift"]

class File(Model):
    model_str = "file"
    single = "File"
    plural = single + 's'
    has_order = False

    name = CharField(max_length=512)
    time_stamp = DateTimeField(auto_now_add=True, null=True,blank=True)
    last_update = DateTimeField(null=True,blank=True)
    document = FileField(upload_to="files/", blank=True, null=True)
    url = URLField(blank=True, null=True)
    type = CharField(max_length=100, blank=True, null=True, choices=get_choices(file_types))

    def __str__(self): return f"{self.name}"
    # def html_patient(self): return self.html("patient")
    def html(self):
        df = pd.read_excel(self.document)
        return df.to_html(classes=['table', 'table-striped', 'table-center'], index=True, justify='left')

    def delete(self, *args, **kwargs):
        self.document.delete()
        super().delete(*args, **kwargs)

nav_models_staff = [Patient, Staff]
nav_models_admin = nav_models_staff + [Job, Day, Month, SMS]
# nav_models_setup = [AvailableShift, JobType, Frequency, InfoCategory, InfoField, PreferredShift, File, Shift, Note, Info]
nav_models_setup = [AvailableShift, JobType, MedicationType, Frequency, InfoCategory, InfoField, File, RecurringJob, Shift, Note, Info]

other = [PreferredShift, Medication]
models = nav_models_admin + nav_models_setup

# model_strs = []
# for model in models: model_strs.append(model.model_str)

# model_strs = ["patient", "staff", "jobtype", "job", "frequency"]
